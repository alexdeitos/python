from datetime import date

from openpyxl import load_workbook


#from DescontaoLoris.valida_cpf import isCpfValid

def __main__():

    file = input(str("Digite o nome do arquivo: "))
    wb = load_workbook(f"{file}.xlsx")
    ws = wb.active

    nome_salvo = input(str(f"Digite o nome do arquivo para salvar: "))
    arquivo = open(f'{nome_salvo}_{date.today()}.txt', 'w+')
    #cpfInvalidos = open(f'cpf_invalidos{date.today()}.txt', 'w+')
    sheet_obj = wb.active

    lines = sheet_obj.max_row
    columns = sheet_obj.max_column

    print(f'numero de linhas :{lines}')
    print(f'numero de colunas: {columns}')

    for l in range(1, lines+1): 
        if l > 1:
            nome = ws.cell(row=l, column=1).value
            #cpf = ws.cell(row=l, column=2).value
            #if not(isCpfValid(cpf)):
            #    cpfInvalidos.write(f"{cpf}\n")
            endereco = ws.cell(row=l, column=2).value
            num_endereco = ws.cell(row=l, column=3).value
            cidade = ws.cell(row=l, column=4).value
            bairro = ws.cell(row=l, column=5).value
            #limite_credito = ws.cell(row=l, column=7).value
            #mensagem = ws.cell(row=l, column=8).value
            sep = ";;"
            arquivo.write(f";{nome}{sep}{endereco}{sep}{num_endereco}{sep}{cidade}{sep}{bairro};\n")
    arquivo.close()
    print(f'Arquivo {nome_salvo}.txt salvo com sucesso!')
    """
    for l in range(2,lines):
        nome = ws.cell(row=l, column=1).value
        cpf  = ws.cell(row=l, column=2).value
        endereco = ws.cell(row=l, column=3).value
        numero = ws.cell(row=l, column=4).value
        cidade = ws.cell(row=l, column=5).value
        bairro = ws.cell(row=l, column=6).value
        limite_credito = ws.cell(row=l, column=7).value
        mensagem = ws.cell(row=l, column=8).value
        sep = ';;'
        arquivo.write(f';{nome}{sep}{cpf}{sep}{endereco}{sep}{numero}{sep}{cidade}{sep}{bairro}{sep}{limite_credito}{sep}'
                      f'{mensagem};\n')

    print(f'sanvaldo o arquivo estrelaBrasil{date.today()}.txt')
    arquivo.close()
    cpfInvalidos.close()
    print(f'Verifique o arquivo cpf_invalidos{date.today()}.txt e corrija os CPF\'s!')
    """
if __name__ == '__main__':
    __main__()
