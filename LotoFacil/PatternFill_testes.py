from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side


def formata_resultados():
    global i
    wb = load_workbook('loto_facil_asloterias_ate_concurso_1883_sorteio.xlsx')
    ws = wb.active
    jogos = list()
    vetJogos = list()
    vetFinal = list()
    ciclo = list()
    vetTemp = list()
    vetTempFinal = list()
    contCiclo = 0
    cont2 = 0
    sep = [5, 10, 15, 20]
    vetJogo = ['x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x',
               'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x']
    letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
               'U', 'V', 'W','X','Y','Z', 'AA', 'AB', 'AC']

    lines = ws.max_row + 1
    columns = ws.max_column + 1
    print(f'linhas :{lines - 1}\nColunas:{columns - 1}')

    yellow = PatternFill(
        start_color='ffff00',
        end_color='000000',
        fill_type='lightVertical'
    )
    black = PatternFill(
        start_color='000000',
        end_color='000000',
        fill_type='solid'
    )
    red = PatternFill(
        start_color='cd5c5c',
        end_color='cd5c5c',
        fill_type='solid'
    )
    ligth_gray = PatternFill(
        start_color='808080',
        end_color='808080',
        fill_type='solid'
    )
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    fonte = Font(
        bold=True,
        size=14,
        italic=False
    )
    alinhamento = Alignment(
        horizontal='center'
    )

    for li in range(1, lines):
        for col in range(1, columns):
            if li > 7 and col > 2:
                i = ws.cell(row=li, column=col).value
                jogos.append(f'{int(i)}')
        vetJogos.append(jogos[:])
        jogos.clear()

    for jogo in vetJogos:
        for j in jogo:
            cont = int(j)
            vetJogo[cont-1] = int(j)
        vetFinal.append(vetJogo[:])
        vetJogo.clear()
        vetJogo = ['x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x',
                   'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x']

    for jogo in vetFinal:
        for j in jogo:
            if cont2 in sep:
                vetTemp.append('||')
            vetTemp.append(j)
            cont2 += 1
        cont2 = 0
        vetTempFinal.append(vetTemp[:])
        vetTemp.clear()
    for line in range(0, len(vetTempFinal)):
        cont1 = 0
        for n in vetTempFinal[line]:
            cont1 += 1
            celula = ws.cell(row=line+1, column=cont1)
            if n == 'x':
                celula.fill = yellow
            elif n == '||':
                celula.fill = black
            else:
                celula.fill = ligth_gray
                if n not in ciclo:
                    ciclo.append(n)
            celula.value = n
            celula.border = thin_border
            celula.font = fonte
            celula.alignment = alinhamento
        for letra in letters:
            ws.column_dimensions[str(letra)].width = 4
        if len(ciclo) == 25:
            contCiclo += 1
            for i in range(1,30):
                if ws.cell(row=line + 1, column=i).value == '||':
                    ws.cell(row=line + 1, column=i).fill = black
                else:
                    ws.cell(row=line + 1, column=i).fill = red
            ws.cell(row=line + 1, column=i + 1).value = f'Ciclo: {contCiclo}'
            ws.cell(row=line + 1, column=i + 1).font = fonte
            ciclo.clear()
    for x in range(0,7):
        ws.delete_rows(1)
    wb.save('loto_facil_formatado.xlsx')
    print(f'Arquivo salvo com sucesso!')
formata_resultados()