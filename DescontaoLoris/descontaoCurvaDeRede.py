from openpyxl import load_workbook
from datetime import date


def __main__():
    print('Lendo arquivo...\n')
    arquivo = input(f'Digite o nome do arquivo a ser lido: ')
    wb = load_workbook(f'{arquivo}.xlsx')
    ws = wb.active

    #SQLs para serem integrados ao script
    sql = 'REPLACE into ligaux values ('
    sqlA = ',2,1002,0,0);'
    sqlB = ',2,1003,0,0);'
    sqlC = ',2,1004,0,0);'
    sqlD = ',2,1005,0,0);'
    sqlNovo = ',2,1006,0,0);'
    sqlReativado = ',2,1007,0,0);'
    sqlQueima = ',2,1008,0,0);'
    sqlInserReplic = 'INSERT INTO replic.querygerada VALUES (NULL,(SELECT b.idmyfg FROM softpharma.scfemp a ' \
                     'INNER JOIN softpharma.myfilial b ON (a.scf_fg = b.idmyfg)),\"'
    sqlInserReplicEnd = '\",NOW(),1,\'N\');'
    sqlDelete = 'DELETE FROM ligaux  WHERE aux_codigo = '
    sqlDeleteMidle = ' AND aux_tipo = 2 AND aux_auxiliar <> '
    sqlEndScript = ';'

    sheet_obj = wb.active
    lines = sheet_obj.max_row + 1

    hoje = date.today()
    curva_nova = open(f'curva{hoje}.sql', 'a')
    curva_replic = open(f'curva_replic{hoje}.sql', 'a')
    delete_curva = open(f'delete_curva{hoje}.sql', 'a')
    delete_curva_replic = open(f'delete_curva_replic{hoje}.sql', 'a')

    print('Montando scripts...\n')
    for l in range(1, lines):

        cod_produto = ws.cell(row=l, column=1)
        curva = ws.cell(row=l, column=2)

        if curva.value == 'A':
            curva_nova.write(f'{sql}{cod_produto.value:.0f}{sqlA}\n')
            curva_replic.write(f'{sqlInserReplic}{sql}{cod_produto.value:.0f}{sqlA}{sqlInserReplicEnd}\n')
            delete_curva.write(f'{sqlDelete}{cod_produto.value:.0f}{sqlDeleteMidle}1002{sqlEndScript}\n')
            delete_curva_replic.write(f'{sqlInserReplic}{sqlDelete}{cod_produto.value:.0f}{sqlDeleteMidle}1002'
                                        f'{sqlEndScript}{sqlInserReplicEnd}\n')

        if curva.value == 'B':
            curva_nova.write(f'{sql}{cod_produto.value:.0f}{sqlB}\n')
            curva_replic.write(f'{sqlInserReplic}{sql}{cod_produto.value:.0f}{sqlB}{sqlInserReplicEnd}\n')
            delete_curva.write(f'{sqlDelete}{cod_produto.value:.0f}{sqlDeleteMidle}1003{sqlEndScript}\n')
            delete_curva_replic.write(f'{sqlInserReplic}{sqlDelete}{cod_produto.value:.0f}{sqlDeleteMidle}1003'
                                      f'{sqlEndScript}{sqlInserReplicEnd}\n')
        if curva.value == 'C':
            curva_nova.write(f'{sql}{cod_produto.value:.0f}{sqlC}\n')
            curva_replic.write(f'{sqlInserReplic}{sql}{cod_produto.value:.0f}{sqlC}{sqlInserReplicEnd}\n')
            delete_curva.write(f'{sqlDelete}{cod_produto.value:.0f}{sqlDeleteMidle}1004{sqlEndScript}\n')
            delete_curva_replic.write(f'{sqlInserReplic}{sqlDelete}{cod_produto.value:.0f}{sqlDeleteMidle}1004'
                                      f'{sqlEndScript}{sqlInserReplicEnd}\n')

        if curva.value == 'D':
            curva_nova.write(f'{sql}{cod_produto.value:.0f}{sqlD}\n')
            curva_replic.write(f'{sqlInserReplic}{sql}{cod_produto.value:.0f}{sqlD}{sqlInserReplicEnd}\n')
            delete_curva.write(f'{sqlDelete}{cod_produto.value:.0f}{sqlDeleteMidle}1005{sqlEndScript}\n')
            delete_curva_replic.write(f'{sqlInserReplic}{sqlDelete}{cod_produto.value:.0f}{sqlDeleteMidle}1005'
                                      f'{sqlEndScript}{sqlInserReplicEnd}\n')

        if curva.value == 'Novo':
            curva_nova.write(f'{sql}{cod_produto.value:.0f}{sqlNovo}\n')
            curva_replic.write(f'{sqlInserReplic}{sql}{cod_produto.value:.0f}{sqlNovo}{sqlInserReplicEnd}\n')
            delete_curva.write(f'{sqlDelete}{cod_produto.value:.0f}{sqlDeleteMidle}1006{sqlEndScript}\n')
            delete_curva_replic.write(f'{sqlInserReplic}{sqlDelete}{cod_produto.value:.0f}{sqlDeleteMidle}1006'
                                      f'{sqlEndScript}{sqlInserReplicEnd}\n')

        if curva.value == 'Reativado':
            curva_nova.write(f'{sql}{cod_produto.value:.0f}{sqlReativado}\n')
            curva_replic.write(f'{sqlInserReplic}{sql}{cod_produto.value:.0f}{sqlReativado}{sqlInserReplicEnd}\n')
            delete_curva.write(f'{sqlDelete}{cod_produto.value:.0f}{sqlDeleteMidle}1007{sqlEndScript}\n')
            delete_curva_replic.write(f'{sqlInserReplic}{sqlDelete}{cod_produto.value:.0f}{sqlDeleteMidle}1007'
                                      f'{sqlEndScript}{sqlInserReplicEnd}\n')

        if curva.value == 'Queima':
            curva_nova.write(f'{sql}{cod_produto.value:.0f}{sqlQueima}\n')
            curva_replic.write(f'{sqlInserReplic}{sql}{cod_produto.value:.0f}{sqlQueima}{sqlInserReplicEnd}\n')
            delete_curva.write(f'{sqlDelete}{cod_produto.value:.0f}{sqlDeleteMidle}1008{sqlEndScript}\n')
            delete_curva_replic.write(f'{sqlInserReplic}{sqlDelete}{cod_produto.value:.0f}{sqlDeleteMidle}1008'
                                      f'{sqlEndScript}{sqlInserReplicEnd}\n')

    print('Arquivo salvo com sucesso!')
    curva_nova.close()
    curva_replic.close()
    delete_curva.close()
    delete_curva_replic.close()


if __name__ == '__main__':
    __main__()