from datetime import date

from openpyxl import load_workbook


def __main__():
    print('Lendo arquivo...\n')
    arquivo = input(f'Digite o nome do arquivo: ')
    wb = load_workbook(f'{arquivo}.xlsx')
    ws = wb.active
    giro = open(f'giro_diario{date.today()}.sql', 'w+')
    giro_replic = open(f'giro_diario_replic{date.today()}.sql', 'w+')

    # SQLs para montagem do arquivo
    sql = "update estcad1 set cad_abc_dias_dem ='"
    sql1 = "' where cad_codigo = "
    sql2 = " and cad_fg = "
    sql3 = ";"
    sql4 = "INSERT INTO replic.querygerada VALUES (NULL,(SELECT b.idmyfg FROM softpharma.scfemp a " \
           "INNER JOIN softpharma.myfilial b ON (a.scf_fg = b.idmyfg)),\""
    sql5 = "\",NOW(),1,'N');"

    sheet_obj = wb.active
    lines = sheet_obj.max_row + 1

    print('Montando scripts...\n')
    for l in range(1, lines):
        fg = ws.cell(row=l, column=1)
        cod_produto = ws.cell(row=l, column=2)
        giro_diario = ws.cell(row=l, column=3)
        giro.write(f'{sql}{giro_diario.value:.2f}{sql1}{cod_produto.value}{sql2}{fg.value}{sql3}\n')
        giro_replic.write(f'{sql4}{sql}{giro_diario.value:.2f}{sql1}{cod_produto.value}{sql2}{fg.value}{sql3}{sql5}\n')

    print('Salvando arquivo...\n')
    giro.close()
    giro_replic.close()
    print('Arquivo salvo com sucesso!')


if __name__ == '__main__':
    __main__()
