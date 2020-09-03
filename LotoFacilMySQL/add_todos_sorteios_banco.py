from openpyxl import load_workbook
from LotoFacilMySQL.connect_mysql import conecta


def addTodosSorteios():
    mydb = conecta()

    wb = load_workbook('loto.xlsx')
    ws = wb.active

    lines = ws.max_row + 1
    columns = ws.max_column + 1
    mycursor = mydb.cursor()
    game = []
    jogos=[]
    contator = 0
    for li in range(1, lines):
        for col in range(0, columns):
            if li > 7 and col > 2:
                i = ws.cell(row=li, column=col).value
                game.append(i)
        jogos.append(sorted(game[:]))
        game.clear()
    for j in jogos:
        if len(j) > 0:
            n1 = j[0]
            n2 = j[1]
            n3 = j[2]
            n4 = j[3]
            n5 = j[4]
            n6 = j[5]
            n7 = j[6]
            n8 = j[7]
            n9 = j[8]
            n10 = j[9]
            n11 = j[10]
            n12 = j[11]
            n13 = j[12]
            n14 = j[13]
            n15 = j[14]
            contator += 1
            sql = f'insert into jogos (dezena_1, dezena_2, dezena_3, dezena_4, dezena_5, dezena_6, dezena_7, ' \
                  f'dezena_8, dezena_9, dezena_10, dezena_11, dezena_12, dezena_13, dezena_14, dezena_15)' \
                  f'values({n1},{n2},{n3},{n4},{n5},{n6},{n7},{n8},{n9},{n10},{n11},{n12},{n13},{n14},{n15})'
            mycursor.execute(sql)
    print(f'Foram adcionados {contator} jogos no banco de dados')
    mydb.commit()
    mydb.close()











