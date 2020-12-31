from openpyxl import load_workbook

from LotoFacilMySQL.connect_mysql import conecta


def desdobra():
    wb = load_workbook(f"desdobra.xlsx")
    ws = wb.active

    lines = ws.max_row
    columns = ws.max_column

    #vetores
    vetorPrimos = [2, 3, 5, 7, 11, 13, 17, 19, 23]
    vetorMoldura = [1, 2, 3, 4, 5, 6, 10, 11, 15, 16, 20, 21, 22, 23, 24, 25]
    vetRepetidas = [3,4,6,7,11,14,15,16,19,20,21,22,23,24,25]
    vetImpar = [1,3,5,7,9,11,13,15,17,19,21,23,25]

    file = open("desdobra.txt" , "w+")
    print(lines)
    print(columns)
    vet1 = []
    vet2 = []
    vet3 = []
    # CRIA O VETOR DO ÚLTIMO RESULTADO PARA MONTAR O VETOR REPETIDAS
    mydb = conecta()
    mycursor = mydb.cursor()
    mycursor1 = mydb.cursor()
    mycursor1.execute("select max(id_concurso) from jogos")
    last = mycursor1.fetchone()
    last = last[0]
    sql = f"SELECT dezena_1, dezena_2, dezena_3, dezena_4, dezena_5, dezena_6, dezena_7, dezena_8,dezena_9, dezena_10, " \
          f"dezena_11, dezena_12, dezena_13, dezena_14, dezena_15 FROM jogos WHERE id_concurso = {last}"
    mycursor.execute(sql)
    result = mycursor.fetchall()
    print("Dezenas sorteadas no concurso anterior: ",end=' ')
    print(result[0])
    ##### fim vetor repetidas

    # define os parametros da validação
    qtd_jogos = int(input("Quantos jogos: "))
    minImpar  = int(input("Minimo Impar: "))
    maxImpar  = int(input("Maximo Impar: "))
    minRep    = int(input("Minimo Repetidas: "))
    maxRep    = int(input("Maximo Repetidas: "))
    minPrimo  = int(input("Minimo Primo: "))
    maxPrimo  = int(input("Maximo Primo: "))
    # maxMold  = int(input("Minimo Moldura: "))
    # minMold  = int(input("Maximo Moldura: "))
    # minFibo = int(input("Minimo Fibo: "))
    # maxFibo = int(input("Maximo Fibo: "))
    # menQtz = int(input("Máximo de dezenas menores que 14: "))

    for l in range(1, lines+1):
        for c in range(1, columns+1):
            n = ws.cell(row=l, column=c).value
            vet1.append(n)
        vet2.append(vet1[:])
        vet1.clear()
    contJogos = 0
    while True:
        for v in vet2:
            # CONTADORES
            contPrimo = contMoldura = contRepetidas = contImpar = soma = 0
            for n in v:
                if n in vetImpar:
                    contImpar += 1
                if n in vetorPrimos:
                    contPrimo += 1
                if n in result[0]:
                    contRepetidas += 1
                if n in vetorMoldura:
                    contMoldura += 1
                soma += n
            if minImpar <= contImpar <=maxImpar and minPrimo <= contPrimo <= maxPrimo and minRep <= contRepetidas <= maxRep\
                    and 180 <= soma <= 215 :
                file.write(f"{sorted(v)}\n")
                file.write(f"impar: {contImpar}\n")
                file.write(f"primo: {contPrimo}\n")
                file.write(f"soma: {soma}\n")
                file.write(f"repetidas: {contRepetidas}\n")

                print(sorted(v))
                print(F"impar: {contImpar}")
                print(F"primo: {contPrimo}")
                print(f"Repetidas: {contRepetidas}")
                print(f"soma:{soma}\n")
                if v not in vet3:
                    vet3.append(v)
                if len(vet3) < qtd_jogos:
                    break
        file.close()
desdobra()
