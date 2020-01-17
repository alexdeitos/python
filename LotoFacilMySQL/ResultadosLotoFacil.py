from openpyxl import *


def resultadosLotofacil():
    #arquivo = input(str('Digite o nome do arquivo: ')).strip().lower()
    wb = load_workbook(f'loto.xlsx')
    ws = wb.active

    linhas = ws.max_row
    colunas = ws.max_column

    vetNum = list()

    print(f'Número de linhas: {linhas}')
    print(f'Número de colunas: {colunas}\n')
    print('#' * 80)
    print()

    for l in range(0, linhas):
        l += 1
        for c in range(0, colunas):
            c+=1
            valor_celula = ws.cell(row=l, column=c).value
            if valor_celula is not None:
                if l <= 6 and c == 1:
                    print(f'{valor_celula:^85}', end=' ')
                elif l == 7:
                    print(f'{valor_celula:^10}', end=' ')
                elif l > 7:
                    if c > 2:
                        print(f'{valor_celula:^10}', end=' ')
                        vetNum.append(valor_celula)
                    else:
                        print(f'{valor_celula:^10}', end=' ')
        print()
resultadosLotofacil()