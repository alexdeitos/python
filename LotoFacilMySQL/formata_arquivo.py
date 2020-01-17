from LotoFacilMySQL.cores_celulas import *
from LotoFacilMySQL.connect_mysql import conecta
from openpyxl import Workbook
from time import sleep


def formata():
    print('Conectando no banco de dados...')
    sleep(2)
    mydb = conecta()

    #Formatações
    yellow = yellow1()
    black = black1()
    red = red1()
    ligth_gray = ligth_gray1()
    thin_border = thin_border1()
    alinhamento = alinhamento1()
    fonte = fonte1()

    wb = Workbook()
    ws = wb.active

    mycursor = mydb.cursor()

    print('Fazendo consulta na tabela jogos...')
    sleep(2)
    mycursor.execute("SELECT dezena_1, dezena_2, dezena_3, dezena_4, dezena_5, dezena_6, dezena_7, dezena_8,"
                     "dezena_9, dezena_10, dezena_11, dezena_12, dezena_13, dezena_14, dezena_15 FROM jogos")
    result = mycursor.fetchall()

    # declaração de variaveis e vetores
    global i
    letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
               'U', 'V', 'W','X','Y','Z', 'AA', 'AB', 'AC']
    vetJogo = ['x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x',
               'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x']
    sep = [5, 10, 15, 20]
    vetJogos = list()
    vetJogosFinal = list()
    ciclo = list()
    contCiclo = 0
    cont2 = 0
    vetTemp = list()

    #inicicio laços de repetição
    for vet in result:
        for num in vet:
            vetJogo[num-1] = num
        vetJogos.append(vetJogo[:])
        vetJogo.clear()
        vetJogo = ['x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x',
                   'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x', 'x']
    for jogo in vetJogos:
        for j in jogo:
            if cont2 in sep:
                vetTemp.append('||')
            vetTemp.append(j)
            cont2 += 1
        cont2 = 0
        vetJogosFinal.append(vetTemp[:])
        vetTemp.clear()
    print('Formatando o arquivo para análise...')
    sleep(1)
    for line in range(0, len(vetJogosFinal)):
        cont1 = 0
        for n in vetJogosFinal[line]:
            cont1 += 1
            celula = ws.cell(row=line+1, column=cont1)
            if n == 'x':
                celula.fill = yellow
            elif n == '||':
                celula.fill = black
            else:
                celula.fill = ligth_gray
                if n not in ciclo:
                    ciclo.append(n)
            celula.value = n
            celula.border = thin_border
            celula.font = fonte
            celula.alignment = alinhamento
        for letra in letters:
            ws.column_dimensions[str(letra)].width = 4
        if len(ciclo) == 25:
            contCiclo += 1
            for i in range(1,30):
                if ws.cell(row=line + 1, column=i).value == '||':
                    ws.cell(row=line + 1, column=i).fill = black
                else:
                    ws.cell(row=line + 1, column=i).fill = red
            ws.cell(row=line + 1, column=i + 1).value = f'Ciclo: {contCiclo}'
            ws.cell(row=line + 1, column=i + 1).font = fonte
            ciclo.clear()
    print('Salvando o arquivo...')
    try:
        wb.save('loto_facil_formatado.xlsx')
    except PermissionError:
        print('\n\n\33[31m O arquivo está aberto, não foi possível salvar o mesmo! Feche o mesmo e tente novamente!'
              '\033[0m\n\n')
    else:
        print(f'\n\n\33[33mArquivo salvo com sucesso! Arquivo loto_facil_formatado.xlsx criado/atualizado'
              f' com sucesso!\33[0m\n\n')
    mydb.close()
    print('Voltando ao Menu Principal...')
    sleep(1.5)
    print('=-=-=-=' * 5)
    print()